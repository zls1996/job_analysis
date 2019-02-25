from openpyxl import Workbook
from openpyxl import load_workbook
import os

#create dir if not exists
def create_dir(dir_path):
    if not os.path.exists(dir_path):
        os.makedirs(dir_path)

def create_file(filename):
    if not os.path.exists(filename):
        with open(filename,'w+') as file:
            file.seek(0)
            file.truncate()
            file.close()

def write_xlsx_header(filename,header):
    wb=Workbook()
    ws1=wb.active
    ws1.title="Sheet1"
    ws1.append(header)
    wb.save(filename)

def write_xlsx_data(filename,data_list):
    wb=load_workbook(filename)
    sheet=wb.get_sheet_by_name("Sheet1")
    for data in data_list:
        sheet.append(data)
    wb.save(filename)