from selenium import webdriver
import re
import time
import csv
import selenium.webdriver.support.expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
import lxml.html
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.chrome.options import Options
etree=lxml.html.etree
import time
import io  
import sys
import random
import os

from openpyxl import Workbook
from openpyxl import load_workbook
#heads=['公司','职位','薪资',
#    '地址','工作经验','学历','职位描述']
heads=['company','position_name','salary',
    'address','work_years','education','job_desp']
path='D:\Software\Google\Chrome\Application\chromedriver.exe'
chrome_options=Options()
#chrome_options.add_argument('--headless')
#chrome_options.add_argument('--proxy-server=http://119.101.114.166:9999')
#driver=webdriver.Chrome(chrome_options=chrome_options,executable_path=path)
driver=webdriver.Chrome(executable_path=path)
proxy_pool=['119.101.114.166:9999','61.135.217.7:80']
count=0
class LagouSpider:
    def __init__(self,driver):
        self.browser=driver
        self.job_kinds_urls=[]
        self.log_filename='log_lagou.txt'
        if not os.path.exists(self.log_filename):
            with open("log_lagou.txt",'w+') as file:
                 file.write(str('0'))
                 file.close()

    def get_all_jobs_kinds(self):
        self.browser.get("https://www.lagou.com")
        try:
            locationAll=self.browser.find_element_by_link_text("全国站")
            locationAll.click()
            WebDriverWait(driver=self.browser,timeout=10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR,
                ".mainNavs")))
        except:
            print("error occured")
            return
        menuBoxs=self.browser.find_elements_by_class_name("menu_sub")
        for menuBox in menuBoxs:
            links=menuBox.find_elements_by_tag_name("a")
            for link in links:
                if link:
                    url_item=[]
                    job_url=link.get_attribute("href")
                    job_kind=link.get_attribute("text")
                    url_item.append(job_kind)
                    url_item.append(job_url)
                    self.job_kinds_urls.append(url_item)

    def print_jobs_kinds(self):
        print(self.job_kinds_urls)

    def request_detail_page(self, url,positions):
        #这里需要修改
        self.browser.execute_script('window.open("%s")'%url)
        self.browser.switch_to.window(self.browser.window_handles[2])
        try:
            WebDriverWait(driver=self.browser,timeout=10).until(EC.presence_of_element_located((By.XPATH,"//span[@class='name']")))
            source = self.browser.page_source
            self.get_detail(source,positions)
        except TimeoutException:
            print("timeout")
        #爬取完这个职位详情页的时候，那么 需要关闭这个详情页，返回列表页
        self.browser.close()
        self.browser.switch_to.window(self.browser.window_handles[1])

    def parse_list_page(self,source,positions):
        html=etree.HTML(source)
        links=html.xpath('//a[@class="position_link"]/@href')
        for link in links:
            self.request_detail_page(link,positions)
            time.sleep(random.randint(2,5))

    def get_detail(self, source,positions):
        job=[]
        html = etree.HTML(source)
        try:
        # 获取页面的内容
            company_name = html.xpath("//img[@class='b2']/@alt")
            job_name = html.xpath("//span[@class='name']/text()")[0]
            job_name=re.sub(r"[\s/]", '', job_name).strip()
            job_requests = html.xpath("//dd[@class='job_request']//span")
            salary = job_requests[0].xpath(".//text()")[0]
            salary = re.sub(r"[\s/]", '', salary).strip()
            address = job_requests[1].xpath(".//text()")[0]
            address = re.sub(r"[\s/]", '', address).strip()
            work_years = job_requests[2].xpath(".//text()")[0]
            work_years = re.sub(r"[\s/]", '', work_years).strip()
            education = job_requests[3].xpath(".//text()")[0]
            education = re.sub(r"[\s/]", '', education).strip()
            job_desp = html.xpath("//dd[@class='job_bt']//text()")
            job_desp = "".join(job_desp)
            job_desp=re.sub(r"[\s/]", '', job_desp).strip()

            job.append(company_name[0])
            job.append(job_name)
            job.append(salary)
            job.append(address)
            job.append(work_years)
            job.append(education)
            job.append(job_desp)
            #list_item_to_gbk(job)
            #print(job)
            positions.append(job)
        except:
            print("failed to get one detail")

    def crawl_url(self,kind,url,positions):
        self.browser.execute_script('window.open("%s")'%url)
        self.browser.switch_to.window(self.browser.window_handles[1])
        kind=kind.replace('/','_')
        filename="./data/"+kind+".xlsx"
        #self.write_file_header(filename)
        self.write_xlsx_header(filename)
        while True:
            source = self.browser.page_source
            self.parse_list_page(source,positions)
            #self.append_to_file(filename,positions)
            self.write_xlsx_data(filename,positions)
            positions=[]
            try:
                #这个next_btn标签可能还没有出现，你是不能点击的，那么我们就是需要显式等待
                WebDriverWait(driver=self.browser,timeout=10).until(EC.presence_of_element_located((By.XPATH,"//div[@class='pager_container']")))
                next_btn = self.browser.find_element_by_link_text("下一页")
                if 'pager_next_disabled' in next_btn.get_attribute('class'):
                    break
                else:
                    next_btn.click()
                    time.sleep(1)
            except:
                print('get next page button failed')
                break
        print("finished 1 url!")       
        self.browser.close()
        self.browser.switch_to.window(self.browser.window_handles[0])

    def write_file_header(self,filename):
        with open(filename,'w',encoding='utf-8',newline='') as fp:
            writer=csv.DictWriter(fp,heads)
            writer.writeheader()
            fp.close()

    def write_xlsx_header(self,filename):
        wb=Workbook()
        ws1=wb.active
        ws1.title="Sheet1"
        ws1.append(heads)
        wb.save(filename)

    def write_xlsx_data(self,filename,data_list):
        wb=load_workbook(filename)
        sheet=wb.get_sheet_by_name("Sheet1")
        for data in data_list:
            sheet.append(data)
        wb.save(filename)


    def append_to_file(self,filename,positions):
         with open(filename,'a',encoding='utf-8',newline='') as fp:
            writer=csv.writer(fp)
            for data in positions:
                writer.writerow(data)

    def crawl_urls(self):
        ready=self.get_log()
        for i in range(0,ready):
            del self.job_kinds_urls[0]
        for url_item in self.job_kinds_urls:
            positions=[]
            self.crawl_url(url_item[0],url_item[1],positions)
            time.sleep(random.randint(40,60))
            self.update_log();

    def update_log(self):
        with open(self.log_filename,'r+') as file:
            data=file.read()
            if not data:
                data='0'
            i=int(data)
            i+=1
            file.seek(0)
            file.truncate()
            file.write(str(i))
            file.close()

    def get_log(self):
        with open(self.log_filename,'r') as file:
            data=file.read()
            if not data:
                return 0
            i=int(data)
            file.close()
            return i

def get_jobs_urls():
    urls=[]
    driver.get("https://www.lagou.com")
    try:
        locationAll=driver.find_element_by_link_text("全国站")
        locationAll.click()
        WebDriverWait(driver=driver,timeout=10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR,
            ".mainNavs")))
    except:
        print("error occured")
    menuBoxs=driver.find_elements_by_class_name("menu_sub")
    for menuBox in menuBoxs:
        links=menuBox.find_elements_by_tag_name("a")
        for link in links:
            if link:
                url_item=[]
                job_url=link.get_attribute("href")
                job_kind=link.get_attribute("text")
                url_item.append(job_kind)
                url_item.append(job_url)
                urls.append(url_item)
    driver.quit()
    return urls
def get_proxy():
    index=count%2
    re=proxy_pool[index]
    count=count+1
    return re
def crawl_lagou(urls):
    for url_item in urls:
        positions=[]
        chrome_options=Options()
        #chrome_options.add_argument('--headless')
        proxy_ip=get_proxy()
        chrome_options.add_argument('--proxy-server=http://'+proxy_ip)
        browser_in_proxy=webdriver.Chrome(chrome_options=chrome_options,
            executable_path=path)
        lagouSpider=LagouSpider(browser_in_proxy)
        lagouSpider.crawl_url(url_item[0],url_item[1],positions)   

if __name__=='__main__': 
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer,encoding='utf8')
    '''
    urls=get_jobs_urls()
    crawl_lagou(urls)
    '''
    
    lagouSpider=LagouSpider(driver)
    lagouSpider.get_all_jobs_kinds()
    lagouSpider.crawl_urls()
    #driver.get("https://www.lagou.com")
    
    #driver.get("http://httpbin.org/ip")